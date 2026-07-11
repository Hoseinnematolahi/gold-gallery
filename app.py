import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
import io

# تنظیمات اولیه پروداکت دیزاین برای حس لوکس و VIP گالری طلا
st.set_page_config(
    page_title="پلتفرم هوشمند مدل‌سازی مالی گالری طلا",
    page_icon="👑",
    layout="wide"
)

# تزریق استایل سفارشی برای راست‌چین کردن و تم لوکس (سرمه‌ای و طلایی)
st.markdown("""
    <style>
    .reportview-container { direction: RTL; text-align: right; }
    .stApp { background-color: #F8F9FA; }
    h1, h2, h3 { color: #1B365D; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; }
    .stButton>button {
        background-color: #1B365D; color: #FFFFFF; 
        border-radius: 8px; border: 1px solid #D4AF37;
        font-weight: bold; width: 100%;
    }
    .stButton>button:hover { background-color: #D4AF37; color: #1B365D; border: 1px solid #1B365D; }
    div[data-testid="stMetricValue"] { color: #1B365D; font-weight: bold; }
    </style>
    """, unsafe_allow_html=True)

# -----------------------------------------------------------------------------
# سایدبار تعاملی
# -----------------------------------------------------------------------------
st.sidebar.header("⚙️ تنظیمات و متغیرهای زنده بازار")
st.sidebar.subheader("💰 پارامترهای فاندامنتال طلا")

live_gold_price = st.sidebar.number_input("قیمت هر گرم طلای ۱۸ عیار خام (تومان)", min_value=1000000, value=4500000, step=50000)
avg_wage = st.sidebar.slider("میانگین درصد اجرت ساخت بنکداری", min_value=0.0, max_value=0.50, value=0.20, step=0.01, format="%.2f")
gallery_markup = st.sidebar.slider("حاشیه سود ناخالص گالری (مارک‌آپ)", min_value=0.05, max_value=0.40, value=0.24, step=0.01, format="%.2f")

st.sidebar.subheader("📈 پیش‌بینی تورم و رشد سالانه")
inflation_gold = st.sidebar.slider("نرخ تورم سالانه قیمت طلا", min_value=0.0, max_value=1.0, value=0.30, step=0.05)
inflation_rent = st.sidebar.slider("نرخ تورم سالانه اجاره ملک", min_value=0.0, max_value=1.0, value=0.35, step=0.05)
inflation_salary = st.sidebar.slider("نرخ رشد سالانه دستمزدها", min_value=0.0, max_value=1.0, value=0.25, step=0.05)

st.sidebar.subheader("👥 ساختار آورده نقدی شرکاء")
partner_1_cash = st.sidebar.number_input("سرمایه شریک اول - سرمایه‌گذار (تومان)", value=5500000000, step=100000000)
partner_2_cash = st.sidebar.number_input("سرمایه شریک دوم - اجرایی (تومان)", value=1500000000, step=50000000)

total_pool_investment = partner_1_cash + partner_2_cash

# -----------------------------------------------------------------------------
# منطق محاسبات در بک‌اند
# -----------------------------------------------------------------------------
capex_fixed_total = (
    1200000000 + 30000000 + 40000000 + 380000000 + 80000000 + 50000000 +
    120000000 + 80000000 + 90000000 + 40000000 + 20000000 + 25000000 +
    35000000 + 15000000 + 25000000 + 60000000 + 40000000 + 50000000
)

gold_working_capital = total_pool_investment - capex_fixed_total
cost_per_gram_built = live_gold_price * (1 + avg_wage)
initial_gold_weight_grams = gold_working_capital / cost_per_gram_built

daily_traffic = 10
conversion_rate = 0.30
avg_invoice_weight = 5
working_days = 26

monthly_sales_weight_grams = daily_traffic * conversion_rate * avg_invoice_weight * working_days
sell_price_per_gram = cost_per_gram_built * (1 + gallery_markup)
monthly_revenue = monthly_sales_weight_grams * sell_price_per_gram

annual_revenue_y1 = monthly_revenue * 12
annual_cogs_y1 = annual_revenue_y1 / (1 + gallery_markup)
annual_gross_profit_y1 = annual_revenue_y1 - annual_cogs_y1

monthly_opex_fixed = 120000000 + 22000000 + 16000000 + 6000000 + 8000000 + 7000000 + 2000000 + 25000000
monthly_opex_variable = 6000000 + 4000000 + 5000000
total_monthly_opex = monthly_opex_fixed + monthly_opex_variable
annual_opex_y1 = total_monthly_opex * 12

ebitda_y1 = annual_gross_profit_y1 - annual_opex_y1
net_profit_y1 = ebitda_y1 * 0.85
roi_y1 = (net_profit_y1 / total_pool_investment) * 100
payback_period_years = total_pool_investment / net_profit_y1 if net_profit_y1 > 0 else 0

# -----------------------------------------------------------------------------
# رندر داشبورد آنلاین
# -----------------------------------------------------------------------------
st.title("👑 پلتفرم هوشمند مدیریت و مدل‌سازی مالی گالری طلا")
st.subheader("لوکیشن استراتژیک مبنا: شیراز - خیابان عفیف‌آباد")
st.markdown("---")

col1, col2, col3, col4 = st.columns(4)
with col1:
    st.metric("کل نقدینگی پروژه", f"{total_pool_investment:,.0f} تومان")
with col2:
    st.metric("شارژ اولیه انبار طلا", f"{gold_working_capital:,.0f} تومان")
with col3:
    st.metric("وزن طلای موجودی اولیه", f"{initial_gold_weight_grams:,.1f} گرم")
with col4:
    st.metric("پیش‌بینی درآمد سال اول", f"{annual_revenue_y1:,.0f} تومان")

col5, col6, col7, col8 = st.columns(4)
with col5:
    st.metric("سود خالص سال اول (پس از مالیات)", f"{net_profit_y1:,.0f} تومان")
with col6:
    st.metric("نرخ بازگشت سرمایه (ROI)", f"{roi_y1:.1f}%")
with col7:
    st.metric("دوره بازگشت سرمایه", f"{payback_period_years * 12:.1f} ماه")
with col8:
    st.metric("قیمت تمام شده هر گرم ساخته شده", f"{cost_per_gram_built:,.0f} تومان")

st.markdown("---")

# -----------------------------------------------------------------------------
# تابع ساخت اکسل فرمول‌نویسی شده
# -----------------------------------------------------------------------------
def generate_dynamic_excel():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    sheet_names = [
        "داشبورد و ورودی‌های کلیدی", "هزینه‌های سرمایه‌ای و جاری", 
        "مدیریت موجودی و مدل فروش", "صورت مالی ۵ ساله و ارزیابی", 
        "شراکت و مکان‌یابی هوشمند", "نمودارهای عملکرد مالی"
    ]
    
    sheets = {}
    for name in sheet_names:
        ws = wb.create_sheet(title=name)
        ws.sheet_view.rightToLeft = True
        ws.views.sheetView[0].showGridLines = True
        sheets[name] = ws

    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=15, bold=True, color="1B365D")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=12, bold=True, color="1B365D")
    bold_font = Font(name=font_family, size=11, bold=True, color="000000")
    regular_font = Font(name=font_family, size=11, color="333333")

    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    sub_header_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    input_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
    kpi_fill = PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid")
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    thin_line = Side(style='thin', color='D1D5DB')
    thick_bottom = Side(style='double', color='1B365D')
    cell_border = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
    total_border = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thick_bottom)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # شیت ۱
    ws = sheets["داشبورد و ورودی‌های کلیدی"]
    ws.cell(row=1, column=1, value="خلاصه داشبورد مدیریتی و شاخص‌های کلیدی پروژه (KPI)").font = title_font
    ws.append([])
    ws.append(["شاخص کلیدی ارزیابی مالی (KPI)", "مقدار محاسباتی داینامیک", "واحد سنجش", "وضعیت توجیه اقتصادی"])
    ws.append(["کل سرمایه نقدی مورد نیاز طرح", "='هزینه‌های سرمایه‌ای و جاری'!$C$21", "تومان", "مطابق با بودجه مصوب"])
    ws.append(["پیش‌بینی درآمد ناخالص سال اول", "='صورت مالی ۵ ساله و ارزیابی'!$B$3", "تومان", "سناریو پایه مبنا"])
    ws.append(["سود خالص نهایی سال اول گالری", "='صورت مالی ۵ ساله و ارزیابی'!$B$8", "تومان", "پس از کسر مالیات فرضی"])
    ws.append(["نرخ بازگشت سرمایه سال اول (ROI)", "='صورت مالی ۵ ساله و ارزیابی'!$B$11", "درصد", "پتانسیل رشد بالا"])
    ws.append(["دوره بازگشت سرمایه کل پروژه", "='صورت مالی ۵ ساله و ارزیابی'!$B$12", "سال", "خروجی مدل زنده"])
    ws.append(["نقطه سر به سر فروش ماهانه", "='صورت مالی ۵ ساله و ارزیابی'!$B$13", "گرم طلا", "سطح ریسک پایدار"])

    for r in range(3, 10):
        fill = header_fill if r == 3 else kpi_fill
        f = header_font if r == 3 else (bold_font if r in [4, 7] else regular_font)
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.font = f; cell.fill = fill; cell.border = cell_border
            cell.alignment = align_center if r == 3 else align_right
        if r > 3:
            ws.cell(row=r, column=2).number_format = '0.0%' if r == 8 else ('0.0' if r == 9 else '#,##0')

    ws.append([]); ws.append([])
    ws.cell(row=12, column=1, value="جدول متغیرها تزریق شده از پنل آنلاین وب‌سایت").font = section_font
    ws.append(["دسته‌بندی اصلی", "نام متغیر فرآیندی", "مقدار ورودی فعال", "واحد سنجش"])

    inputs_data = [
        ["پارامترهای بازار طلا", "قیمت پایه هر گرم طلای ۱۸ عیار خام", live_gold_price, "تومان"],
        ["پارامترهای بازار طلا", "میانگین اجرت و کارمزد ساخت خرید گالری", avg_wage, "درصد"],
        ["پارامترهای بازار طلا", "حاشیه سود ناخالص گالری (مارک‌آپ خرده‌فروشی)", gallery_markup, "درصد"],
        ["نرخ‌های رشد و تورم", "نرخ رشد سالانه قیمت طلا (تورم طلا)", inflation_gold, "درصد"],
        ["نرخ‌های رشد و تورم", "نرخ تورم سالانه هزینه‌های اجاره ملک", inflation_rent, "درصد"],
        ["نرخ‌های رشد و تورم", "نرخ رشد سالانه دستمزد و حقوق پرسنل", inflation_salary, "درصد"],
        ["نرخ‌های رشد و تورم", "نرخ تورم هزینه‌های عمومی و اقلام متغیر", 0.20, "درصد"],
        ["نرخ‌های رشد و تورم", "نرخ رشد سالانه حجم فیزیکی فروش (تعداد مشتری)", 0.10, "درصد"],
        ["آورده شرکاء گالری", "سرمایه نقدی واریزی شریک اول (سرمایه‌گذار)", partner_1_cash, "تومان"],
        ["آورده شرکاء گالری", "سرمایه نقدی واریزی شریک دوم (اجرایی)", partner_2_cash, "تومان"],
        ["عملیات فروشگاه فیزیکی", "تعداد روزهای کاری ماهانه گالری", working_days, "روز"]
    ]
    for row in inputs_data: ws.append(row)

    for r in range(13, 25):
        fill = header_fill if r == 13 else input_fill
        f = header_font if r == 13 else regular_font
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.font = f; cell.border = cell_border
            cell.alignment = align_center if r == 13 else align_right
            if r > 13 and c == 3: cell.fill = input_fill
        if r > 13:
            if r in range(15, 21): ws.cell(row=r, column=3).number_format = '0.0%'
            elif r in [14, 22, 23]: ws.cell(row=r, column=3).number_format = '#,##0'

    # شیت ۲
    ws = sheets["هزینه‌های سرمایه‌ای و جاری"]
    ws.cell(row=1, column=1, value="تفکیک ریز هزینه‌های راه‌اندازی (CAPEX) و جاری عملیاتی (OPEX)").font = title_font
    ws.append([])
    ws.append(["دسته‌بندی اصلی", "کد هزینه", "ریز عنوان هزینه (Line Item)", "مبلغ اولیه (تومان)", "فرکانس", "توضیحات استراتژیک"])

    capex_items = [
        ["ودیعه ملک (Property)", "CAP-01", "ودیعه (رهن) ملک تجاری در عفیف‌آباد", 1200000000, "یکبار زمان شروع", "مغازه ۱۸ تا ۲۵ متر"],
        ["ودیعه ملک (Property)", "CAP-02", "حق کمیسیون مشاورین املاک و ثبت قرارداد", 30000000, "یکبار زمان شروع", "مخارج قانونی قرارداد"],
        ["طراحی و دکوراسیون", "CAP-03", "طراحی معماری و مهندسی داخلی", 40000000, "یکبار زمان شروع", "کانسپت مدرن هیبرید"],
        ["طراحی و دکوراسیون", "CAP-04", "اجرای دکوراسیون چوب، شیشه و متریال لوکس", 380000000, "یکبار زمان شروع", "ویترین‌های ضدسرقت داخلی"],
        ["طراحی و دکوراسیون", "CAP-05", "سیستم نورپردازی تخصصی SMD و هالوژن‌ها", 80000000, "یکبار زمان شروع", "حفظ درخشش طبیعی طلا"],
        ["طراحی و دکوراسیون", "CAP-06", "تابلوی بیرونی گالری (چلنیوم و کامپوزیت برند)", 50000000, "یکبار زمان شروع", "برندینگ فیزیکی محلی"],
        ["تجهیزات امنیتی الکترونیک", "CAP-07", "گاوصندوق فوق سنگین مکانیکی زیرزمینی", 120000000, "یکبار زمان شروع", "گواهی ایمنی سطح بالا"],
        ["تجهیزات امنیتی الکترونیک", "CAP-08", "سیستم دوربین مداربسته ضدآستگمات AI", 80000000, "یکبار زمان شروع", "تشخیص چهره و زوم بالا"],
        ["تجهیزات امنیتی الکترونیک", "CAP-09", "شیشه ضدگلوله لایه مخفی کلاس ۳ برای ویترین", 90000000, "یکبار زمان شروع", "مقاومت بالا در برابر ضربه"],
        ["تجهیزات امنیتی الکترونیک", "CAP-10", "سیستم مه ساز امنیتی پدالی هوشمند", 40000000, "یکبار زمان شروع", "کورکننده دید سارقین"],
        ["تجهیزات امنیتی الکترونیک", "CAP-11", "دزدگیر مرکزی متصل به سامانه کلانتری (مها)", 20000000, "یکبار زمان شروع", "امنیت یکپارچه گالری"],
        ["زیرساخت فناوری و سخت‌افزار", "CAP-12", "ترازوهای دیجیتال زرگری کالیبره شده دقیق", 25000000, "یکبار زمان شروع", "دارای استاندارد ملی"],
        ["زیرساخت فناوری و سخت‌افزار", "CAP-13", "سیستم صندوق فروشگاهی، بارکدخوان و فیش‌پرینتر", 35000000, "یکبار زمان شروع", "سخت‌افزار پایدار"],
        ["زیرساخت فناوری و سخت‌افزار", "CAP-14", "نرم‌افزار حسابداری تخصصی خرید و فروش طلا", 15000000, "یکبار زمان شروع", "پشتیبانی سالانه"],
        ["زیرساخت فناوری و سخت‌افزار", "CAP-15", "نرم‌افزار CRM و تبلت‌های تعاملی مشتریان", 25000000, "یکبار زمان شروع", "مدیریت ارتباطات لوکس"],
        ["مارکتینگ زمان راه‌اندازی", "CAP-16", "مراسم و ایونت رسمی افتتاحیه لوکس گالری", 60000000, "یکبار زمان شروع", "جذب مشتریان VIP شیراز"],
        ["مارکتینگ زمان راه‌اندازی", "CAP-17", "تولید تیزر تجاری، کاتالوگ و عکاسی صنعتی طلا", 40000000, "یکبار زمان شروع", "محتوای اولیه اینستاگرام"],
        ["مارکتینگ زمان راه‌اندازی", "CAP-18", "کمپین تبلیغاتی با اینفلوئنسرهای لوکس شیراز", 50000000, "یکبار زمان شروع", "ایجاد آگاهی از برند اولیه"],
        ["موجودی اولیه طلا", "CAP-19", "خرید شارژ اولیه انبار طلا (سرمایه در گردش)", "='داشبورد و ورودی‌های کلیدی'!$C$22+'داشبورد و ورودی‌های کلیدی'!$C$23-SUM(D4:D21)", "سرمایه در گردش", "تامین ارگانیک موجودی انبار طلا"]
    ]
    for row in capex_items: ws.append(row)
    ws.append(["جمع کل هزینه‌های سرمایه‌ای ثابت (CAPEX)", "", "", "=SUM(D4:D22)", "", "کل نقدینگی تامین‌شده"])

    for r in range(3, 24):
        fill = header_fill if r == 3 else (total_fill if r == 23 else None)
        f = header_font if r == 3 else (bold_font if r == 23 else regular_font)
        b = total_border if r == 23 else cell_border
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.font = f; cell.fill = fill if fill else PatternFill(fill_type=None); cell.border = b
            cell.alignment = align_center if r == 3 else align_right
        if r > 3: ws.cell(row=r, column=4).number_format = '#,##0'

    ws.append([]); ws.append([])
    ws.cell(row=26, column=1, value="تفکیک ریز هزینه‌های جاری عملیاتی ماهانه (OPEX)").font = section_font
    ws.append(["دسته‌بندی اصلی", "کد هزینه", "ریز عنوان هزینه جاری ماهانه", "مبلغ ماهانه (تومان)", "مبلغ سالانه (تومان)", "نوع هزینه"])

    opex_items = [
        ["هزینه‌های ثابت عملیاتی", "OPX-01", "اجاره‌بهای ماهانه مغازه عفیف‌آباد", 120000000, "=D28*12", "ثابت"],
        ["هزینه‌های ثابت عملیاتی", "OPX-02", "حقوق پایه کارشناس ارشد فروش (مدیر گالری)", 22000000, "=D29*12", "ثابت"],
        ["هزینه‌های ثابت عملیاتی", "OPX-03", "حقوق کارشناس فروش فیزیکی و ادمین آنلاین", 16000000, "=D30*12", "ثابت"],
        ["هزینه‌های ثابت عملیاتی", "OPX-04", "حق‌الزحمه حسابدار مقیم و پکیج خدمات قانونی", 6000000, "=D31*12", "ثابت"],
        ["هزینه‌های ثابت عملیاتی", "OPX-05", "حق بیمه مسئولیت کارفرما و بیمه تمام‌خطر طلا", 8000000, "=D32*12", "ثابت"],
        ["هزینه‌های ثابت عملیاتی", "OPX-06", "شارژ مجتمع تجاری، امنیت پاساژ و شارژها", 7000000, "=D33*12", "ثابت"],
        ["هزینه‌های ثابت عملیاتی", "OPX-07", "هزینه اینترنت پرسرعت، اشتراک نرم‌افزارها", 2000000, "=D34*12", "ثابت"],
        ["هزینه‌های ثابت عملیاتی", "OPX-08", "بودجه مستمر دیجیتال مارکتینگ و تولید ویدیو ماهانه", 25000000, "=D35*12", "ثابت"],
        ["هزینه‌های متغیر عملیاتی", "OPX-09", "جعبه‌های هاردباکس لوکس، شاپینگ بگ بسته‌بندی", 6000000, "=D36*12", "متغیر"],
        ["هزینه‌های متغیر عملیاتی", "OPX-10", "کارمزد پوزهای بانکی و تراکنش‌های مالی", 4000000, "=D37*12", "متغیر"],
        ["هزینه‌های متغیر عملیاتی", "OPX-11", "هزینه پذیرایی VIP، خدمات مشتری و تشریفات", 5000000, "=D38*12", "متغیر"]
    ]
    for row in opex_items: ws.append(row)
    ws.append(["جمع کل هزینه‌های عملیاتی جاری (OPEX)", "", "", "=SUM(D28:D38)", "=SUM(E28:E38)", ""])

    for r in range(27, 41):
        fill = header_fill if r == 27 else (total_fill if r == 40 else None)
        f = header_font if r == 27 else (bold_font if r == 40 else regular_font)
        b = total_border if r == 40 else cell_border
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.font = f; cell.fill = fill if fill else PatternFill(fill_type=None); cell.border = b
            cell.alignment = align_center if r == 27 else align_right
        if r > 27:
            ws.cell(row=r, column=4).number_format = '#,##0'
            ws.cell(row=r, column=5).number_format = '#,##0'

    # شیت ۳
    ws = sheets["مدیریت موجودی و مدل فروش"]
    ws.cell(row=1, column=1, value="مدیریت پویای سبد موجودی اولیه طلا و پیش‌بینی سناریوهای فروش ماهانه").font = title_font
    ws.append([])
    ws.append(["شاخص پایه انبارداری طلا", "فرمول محاسباتی داینامیک", "واحد", "توضیح"])
    ws.append(["کل نقدینگی تخصیص‌یافته به خرید طلا", "='هزینه‌های سرمایه‌ای و جاری'!$D$22", "تومان", "از شیت هزینه تامین می‌شود"])
    ws.append(["قیمت تمام‌شده خرید هر گرم طلای ساخته‌شده برای گالری", "='داشبورد و ورودی‌های کلیدی'!$C$14*(1+'داشبورد و ورودی‌های کلیدی'!$C$15)", "تومان / گرم", "قیمت خام + اجرت ساخت بنکدار"])
    ws.append(["کل وزن طلای اولیه قابل خرید گالری", "=B4/B5", "گرم فیزیکی", "موجودی وزنی اولیه گالری"])

    for r in range(3, 7):
        fill = header_fill if r == 3 else None
        f = header_font if r == 3 else regular_font
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.font = f; cell.fill = fill if fill else PatternFill(fill_type=None); cell.border = cell_border
            cell.alignment = align_center if r == 3 else (Alignment(horizontal="left") if c == 2 else align_right)
        if r > 3: ws.cell(row=r, column=2).number_format = '#,##0.0' if r == 6 else '#,##0'

    ws.append([])
    ws.append(["گروه محصول ویترین طلا", "سهم از سرمایه موجودی", "وزن اختصاص‌یافته (گرم)", "حاشیه سود ناخالص هر دسته", "سرعت گردش دسته در ویترین"])
    portfolio = [
        ["انگشتر طلا (مدل‌های فانتزی و مدرن)", 0.20, "=B$6*B9", 0.25, "بسیار بالا"],
        ["گردنبند، مدال و آویزهای ظریف", 0.22, "=B$6*B10", 0.22, "متوسط رو به بالا"],
        ["دستبندهای نخی، چرمی و النگویی مدرن", 0.18, "=B$6*B11", 0.24, "بالاترین سرعت"],
        ["گوشواره‌های سبک و پیرسینگ طلا", 0.12, "=B$6*B12", 0.26, "بالا"],
        ["سرویس‌های کامل و نیم‌ست‌های لوکس", 0.15, "=B$6*B13", 0.28, "پایین (خواب دار)"],
        ["حلقه‌های ازدواج و ست‌های نامزدی مینیمال", 0.08, "=B$6*B14", 0.22, "متوسط پایداری"],
        ["شمش‌های گوایی‌شده کادویی و لوکس (وزن سبک)", 0.05, "=B$6*B15", 0.07, "سرعت آنی نقدینگی"]
    ]
    for row in portfolio: ws.append(row)
    ws.append(["جمع کل سبد دارایی طلای گالری", "=SUM(B9:B15)", "=SUM(C9:C15)", "='داشبورد و ورودی‌های کلیدی'!$C$16", "میانگین موزون کل"])

    for r in range(8, 17):
        fill = header_fill if r == 8 else (total_fill if r == 16 else None)
        f = header_font if r == 8 else (bold_font if r == 16 else regular_font)
        b = total_border if r == 16 else cell_border
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.font = f; cell.fill = fill if fill else PatternFill(fill_type=None); cell.border = b
            cell.alignment = align_center if r == 8 else align_right
        if r > 8:
            ws.cell(row=r, column=2).number_format = '0.0%'
            ws.cell(row=r, column=3).number_format = '#,##0.0'
            ws.cell(row=r, column=4).number_format = '0.0%'

    ws.append([]); ws.append([])
    ws.cell(row=19, column=1, value="مدل‌سازی پیش‌بینی فروش بر اساس سناریوهای بازار").font = section_font
    ws.append(["شاخص پویای درآمدی ماهانه", "سناریو محافظه‌کارانه", "سناریو پایه (مبنای طرح)", "سناریو خوش‌بینانه"])
    ws.append(["تعداد مشتریان ورودی روزانه به گالری", 5, 10, 20])
    ws.append(["نرخ تبدیل خریدار نهایی (Conversion Rate)", 0.20, 0.30, 0.40])
    ws.append(["میانگین وزن هر فاکتور فروش (گرم)", 3, 5, 8])
    ws.append(["حجم طلای فروخته‌شده در ماه (گرم)", "=B21*B22*B23*'داشبورد و ورودی‌های کلیدی'!$C$24", "=C21*C22*C23*'داشبورد و ورودی‌های کلیدی'!$C$24", "=D21*D22*D23*'داشبورد و ورودی‌های کلیدی'!$C$24"])
    ws.append(["قیمت فروش هر گرم طلا با حاشیه گالری", "='مدیریت موجودی و مدل فروش'!$B$5*(1+'داشبورد و ورودی‌های کلیدی'!$C$16)", "='مدیریت موجودی و مدل فروش'!$B$5*(1+'داشبورد و ورودی‌های کلیدی'!$C$16)", "='مدیریت موجودی و مدل فروش'!$B$5*(1+'داشبورد و ورودی‌های کلیدی'!$C$16)"])
    ws.append(["کل درآمد ریالی فروش ماهانه گالری (تومان)", "=B24*B25", "=C24*C25", "=D24*D25"])
    ws.append(["تعداد ماه‌های لازم برای گردش موجودی", "='مدیریت موجودی و مدل فروش'!$B$6/B24", "='مدیریت موجودی و مدل فروش'!$B$6/C24", "='مدیریت موجودی و مدل فروش'!$B$6/D24"])

    for r in range(20, 28):
        fill = header_fill if r == 20 else (sub_header_fill if r == 26 else None)
        f = header_font if r == 20 else (bold_font if r == 26 else regular_font)
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.font = f; cell.fill = fill if fill else PatternFill(fill_type=None); cell.border = cell_border
            cell.alignment = align_center if r == 20 else align_right
        if r > 20:
            if r == 22:
                for col in [2, 3, 4]: ws.cell(row=r, column=col).number_format = '0.0%'
            elif r in [24, 27]:
                for col in [2, 3, 4]: ws.cell(row=r, column=col).number_format = '#,##0.0'
            elif r in [25, 26]:
                for col in [2, 3, 4]: ws.cell(row=r, column=col).number_format = '#,##0'

    # شیت ۴
    ws = sheets["صورت مالی ۵ ساله و ارزیابی"]
    ws.cell(row=1, column=1, value="صورت سود و زیان جامع ۵ ساله گالری و شاخص‌های مالی ارزیابی طرح").font = title_font
    ws.append([])
    ws.append(["آیتم ساختار سود و زیان (تومان)", "سال ۱", "سال ۲", "سال ۳", "سال ۴", "سال ۵"])

    ws.append(["درآمد ناخالص عملیاتی کل (Revenue)", "='مدیریت موجودی و مدل فروش'!$C$26*12", "=B3*(1+'داشبورد و ورودی‌های کلیدی'!$C$17)*(1+'داشبورد و ورودی‌های کلیدی'!$C$21)", "=C3*(1+'داشبورد و ورودی‌های کلیدی'!$C$17)*(1+'داشبورد و ورودی‌های کلیدی'!$C$21)", "=D3*(1+'داشبورد و ورودی‌های کلیدی'!$C$17)*(1+'داشبورد و ورودی‌های کلیدی'!$C$21)", "=E3*(1+'داشبورد و ورودی‌های کلیدی'!$C$17)*(1+'داشبورد و ورودی‌های کلیدی'!$C$21)"])
    ws.append(["بهای تمام‌شده طلای فروخته‌شده (COGS)", "=B3/(1+'داشبورد و ورودی‌های کلیدی'!$C$16)", "=C3/(1+'داشبورد و ورودی‌های کلیدی'!$C$16)", "=D3/(1+'داشبورد و ورودی‌های کلیدی'!$C$16)", "=E3/(1+'داشبورد و ورودی‌های کلیدی'!$C$16)", "=F3/(1+'داشبورد و ورودی‌های کلیدی'!$C$16)"])
    ws.append(["سود ناخالص عملیاتی گالری (Gross Profit)", "=B3-B4", "=C3-C4", "=D3-D4", "=E3-E4", "=F3-F4"])

    ws.append(["هزینه‌های جاری کل گالری (OPEX)", "='هزینه‌های سرمایه‌ای و جاری'!$E$40", 
               "=('هزینه‌های سرمایه‌ای و جاری'!$E$28*(1+'داشبورد و ورودی‌های کلیدی'!$C$18)) + (SUM('هزینه‌های سرمایه‌ای و جاری'!$E$29:$E$35)*(1+'داشبورد و ورودی‌های کلیدی'!$C$19)) + (SUM('هزینه‌های سرمایه‌ای و جاری'!$E$36:$E$38)*(1+'داشبورد و ورودی‌های کلیدی'!$C$20))",
               "=('هزینه‌های سرمایه‌ای و جاری'!$E$28*(1+'داشبورد و ورودی‌های کلیدی'!$C$18)^2) + (SUM('هزینه‌های سرمایه‌ای و جاری'!$E$29:$E$35)*(1+'داشبورد و ورودی‌های کلیدی'!$C$19)^2) + (SUM('هزینه‌های سرمایه‌ای و جاری'!$E$36:$E$38)*(1+'داشبورد و ورودی‌های کلیدی'!$C$20)^2)",
               "=('هزینه‌های سرمایه‌ای و جاری'!$E$28*(1+'داشبورد و ورودی‌های کلیدی'!$C$18)^3) + (SUM('هزینه‌های سرمایه‌ای و جاری'!$E$29:$E$35)*(1+'داشبورد و ورودی‌های کلیدی'!$C$19)^3) + (SUM('هزینه‌های سرمایه‌ای و جاری'!$E$36:$E$38)*(1+'داشبورد و ورودی‌های کلیدی'!$C$20)^3)",
               "=('هزینه‌های سرمایه‌ای و جاری'!$E$28*(1+'داشبورد و ورودی‌های کلیدی'!$C$18)^4) + (SUM('هزینه‌های سرمایه‌ای و جاری'!$E$29:$E$35)*(1+'داشبورد و ورودی‌های کلیدی'!$C$19)^4) + (SUM('هزینه‌های سرمایه‌ای و جاری'!$E$36:$E$38)*(1+'داشبورد و ورودی‌های کلیدی'!$C$20)^4)"])

    ws.append(["سود عملیاتی قبل از مالیات (EBITDA)", "=B5-B6", "=C5-C6", "=D5-D6", "=E5-E6", "=F5-F6"])
    ws.append(["سود خالص نهایی پس از مالیات فرضی (Net Profit)", "=B7*0.85", "=C7*0.85", "=D7*0.85", "=E7*0.85", "=F7*0.85"])

    for r in range(3, 9):
        fill = header_fill if r == 3 else (total_fill if r == 8 else None)
        f = header_font if r == 3 else (bold_font if r == 8 else regular_font)
        b = total_border if r == 8 else cell_border
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.font = f; cell.fill = fill if fill else PatternFill(fill_type=None); cell.border = b
            cell.alignment = align_center if r == 3 else align_right
            if r > 3 and c > 1: cell.number_format = '#,##0'

    ws.append([]); ws.append([])
    ws.cell(row=11, column=1, value="ارزیابی بازگشت سرمایه و نقطه سر به سر").font = section_font
    ws.append(["شاخص ارزیابی مالی طرح", "مقدار داینامیک", "واحد سنجش / توصیف ریسک"])
    ws.append(["نرخ بازگشت سرمایه سال اول گالری (ROI)", "=B8/'داشبورد و ورودی‌های کلیدی'!$C$4", "درصد پایداری سود"])
    ws.append(["دوره بازگشت سرمایه فیزیکی پروژه", "='هزینه‌های سرمایه‌ای و جاری'!$D$23/B8", "سال شمسی تا نقد شدن اصل آورده"])
    ws.append(["نقطه سر به سر فیزیکی فروش ماهانه گالری", "=('هزینه‌های سرمایه‌ای و جاری'!$D$40)/('مدیریت موجودی و مدل فروش'!$B$5*'داشبورد و ورودی‌های کلیدی'!$C$16)", "گرم طلا در ماه جهت صفر شدن سود"])

    for r in range(12, 16):
        fill = header_fill if r == 12 else None
        f = header_font if r == 12 else regular_font
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.font = f; cell.fill = fill if fill else PatternFill(fill_type=None); cell.border = cell_border
            cell.alignment = align_center if r == 12 else (Alignment(horizontal="left") if c == 2 else align_right)
        if r > 12:
            ws.cell(row=r, column=2).number_format = '0.0%' if r == 13 else ('0.0' if r == 14 else '#,##0.0')

    # شیت ۵
    ws = sheets["شراکت و مکان‌یابی هوشمند"]
    ws.cell(row=1, column=1, value="مدل حقوقی سهامداری شرکاء و ماتریس مکان‌یابی هوشمند شعب شیراز").font = title_font
    ws.append([])
    ws.append(["نام شریک تجاری", "آورده نقدی اولیه (تومان)", "سهم نقدی درصد", "نقش عملیاتی و ارزش‌آفرینی ویژه", "درصد مالکیت سهام نهایی"])
    ws.append(["شریک اول (A - سرمایه‌گذار کلان طرح)", "='داشبورد و ورودی‌های کلیدی'!$C$22", "=B4/'هزینه‌های سرمایه‌ای و جاری'!$D$23", "رئیس هیئت مدیره - ناظر زنجیره تامین مالی", 0.70])
    ws.append(["شریک دوم (B - شریک اجرایی و متخصص)", "='داشبورد و ورودی‌های کلیدی'!$C$23", "=B5/'هزینه‌های سرمایه‌ای و جاری'!$D$23", "مدیرعامل - مقیم گالری - کارشناس طلا و جواهر", 0.30])
    ws.append(["جمع کل ساختار شراکت گالری", "=SUM(B4:B5)", "=SUM(C4:C5)", "-", "=SUM(E4:E5)"])

    for r in range(3, 7):
        fill = header_fill if r == 3 else (total_fill if r == 6 else None)
        f = header_font if r == 3 else (bold_font if r == 6 else regular_font)
        b = total_border if r == 6 else cell_border
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.font = f; cell.fill = fill if fill else PatternFill(fill_type=None); cell.border = b
            cell.alignment = align_center if r == 3 else align_right
        if r > 3:
            ws.cell(row=r, column=2).number_format = '#,##0'
            ws.cell(row=r, column=3).number_format = '0.0%'
            ws.cell(row=r, column=5).number_format = '0.0%'

    ws.append([]); ws.append([])
    ws.cell(row=9, column=1, value="ماتریس امتیازدهی موزون مناطق جغرافیایی شیراز (Location Intelligence)").font = section_font
    ws.append(["منطقه کاندید استقرار", "ترافیک پیاده (۲۰٪)", "کیفیت و قدرت خرید مشتری (۲۵٪)", "سطح رقابت منطقه‌ای (۱۵٪)", "بازدهی و کارایی اجاره (۲۰٪)", "پتانسیل برندینگ گالری (۲۰٪)", "امتیاز نهایی کل (از ۱۰۰)"])

    loc_data = [
        ["خیابان عفیف‌آباد شیراز", 90, 95, 80, 75, 95, "=B11*0.2+C11*0.25+D11*0.15+E11*0.2+F11*0.2"],
        ["بلوار معالی‌آباد", 95, 85, 85, 80, 85, "=B12*0.2+C12*0.25+D12*0.15+E12*0.2+F12*0.2"],
        ["بازار زرگرهای شیراز (سنتی)", 100, 70, 30, 85, 40, "=B13*0.2+C13*0.25+D13*0.15+E13*0.2+F13*0.2"],
        ["خیابان ستارخان", 80, 90, 80, 65, 85, "=B14*0.2+C14*0.25+D14*0.15+E14*0.2+F14*0.2"],
        ["خیابان ملاصدرا", 85, 75, 75, 70, 70, "=B15*0.2+C15*0.25+D15*0.15+E15*0.2+F15*0.2"],
        ["بلوار فرهنگ‌شهر", 60, 95, 90, 60, 90, "=B16*0.2+C16*0.25+D16*0.15+E16*0.2+F16*0.2"]
    ]
    for row in loc_data: ws.append(row)

    for r in range(10, 17):
        fill = header_fill if r == 10 else None
        f = header_font if r == 10 else regular_font
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.font = f; cell.fill = fill if fill else PatternFill(fill_type=None); cell.border = cell_border
            cell.alignment = align_center if r == 10 else align_right
        if r > 10:
            ws.cell(row=r, column=7).font = bold_font
            ws.cell(row=r, column=7).fill = kpi_fill
            ws.cell(row=r, column=7).number_format = '0.0'

    # شیت ۶
    ws = sheets["نمودارهای عملکرد مالی"]
    ws.cell(row=1, column=1, value="تجسم بصری پیش‌بینی‌های مالی ۵ ساله و روند سودآوری گالری").font = title_font
    ws.append([])

    chart = LineChart()
    chart.title = "روند رشد درآمد کل در مقابل سود خالص ۵ ساله گالری طلا"
    chart.style = 13
    chart.y_axis.title = "تومان"
    chart.x_axis.title = "سال مالی"

    data = Reference(sheets["صورت مالی ۵ ساله و ارزیابی"], min_col=2, min_row=3, max_col=6, max_row=3)
    data_net = Reference(sheets["صورت مالی ۵ ساله و ارزیابی"], min_col=2, min_row=8, max_col=6, max_row=8)
    chart.add_data(data, from_rows=True)
    chart.add_data(data_net, from_rows=True)

    cats = Reference(sheets["صورت مالی ۵ ساله و ارزیابی"], min_col=2, min_row=2, max_col=6, max_row=2)
    chart.set_categories(cats)
    ws.add_chart(chart, "B3")

    # فیت کردن ستون‌ها
    for fa_name, sheet in sheets.items():
        for col in sheet.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len: max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 18)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

st.write("### 📥 خروجی نهایی مدل جامع تجاری")
st.info("با کلیک بر روی دکمه زیر، فایل اکسلِ ساخته شده بر اساس مقادیر وارد شده دانلود خواهد شد.")

excel_data = generate_dynamic_excel()
st.download_button(
    label="📊 دانلود مدل مالی فرمول‌نویسی شده (Excel)",
    data=excel_data,
    file_name="Gold_Gallery_Financial_Model.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
